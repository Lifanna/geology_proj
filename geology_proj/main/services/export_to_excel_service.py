import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.cell_style import CellStyle
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from main import models
from openpyxl.utils import get_column_letter
import json, os
from django.conf import settings
from django.db.models import Max


class ExportToExcelService:
    def __init__(self):
        self.workbook = openpyxl.load_workbook(os.path.realpath(os.path.dirname(__file__)) + "/example.xlsx")
        self.worksheet_1 = self.workbook.active
        self.worksheet_2 = self.workbook.get_sheet_by_name('Стр. 2')

    def build_document(self, license : models.License, 
                       watercourse: models.WaterCourse, 
                       watercourse_bound: models.LicenseWaterCourse, 
                       line: models.Line, 
                       well
    ):
            error_message = ""
            filepath = "%s/example.xlsx"%settings.MEDIA_ROOT
        # try:
            self.worksheet_1.title = license.short_name

            self.worksheet_1.cell(column=1, row=1, value=license.short_name)
            self.worksheet_1.cell(column=1, row=1).font = Font(bold=True)

            self.worksheet_1.cell(column=14, row=2, value=well.name)

            # отображаем водотоки
            self.worksheet_1.cell(column=1, row=3, value=watercourse.name)

            # если имеется главный водоток
            if watercourse_bound.parent_watercourse and watercourse_bound.parent_watercourse.id != watercourse_bound.watercourse.id:
                self.worksheet_1.cell(column=6, row=3, value=watercourse_bound.parent_watercourse.name)

            # отображаем линии
            self.worksheet_1.cell(column=1, row=4, value=line.name)
            self.worksheet_1.cell(column=6, row=4, value=well.name)

            # отображаем сведения о скважинах и их интервалах
            # отображаем даты бурения скважин: начала и окончания
            self.worksheet_1.cell(column=3, row=5, value=well.created_at.date())
            self.worksheet_1.cell(column=12, row=5, value=well.updated_at.date())

            layers = models.Layer.objects.filter(well__id=well.id)

            row_index = 9
            cells_index = 9
            prev_depth = 0
            depth_step = 0.5

            n_cells_for_layers = int(layers.aggregate(Max('depth')).get("depth__max") / 0.5)

            for cell_layer_index in range(1, n_cells_for_layers + 1):
                thin_border = Border(left=Side(style='thin', color='000000'), 
                                    right=Side(style='thin', color='000000'), 
                                    top=Side(style='thin', color='000000'), 
                                    bottom=Side(style='thin', color='000000'))
                
                # my_style = CellStyle(border=thin_border)

                self.worksheet_1.cell(column=1, row=cells_index, value=cell_layer_index)

                self.worksheet_1.cell(column=2, row=cells_index, value=str(depth_step))


                self.worksheet_1.cell(column=10, row=cells_index, value=str(6900))

                self.worksheet_1.cell(column=11, row=cells_index, value=str(100))

                self.worksheet_1.cell(column=12, row=cells_index, value=str(6900))

                self.worksheet_1.cell(column=13, row=cells_index, value=str("пс"))

                self.worksheet_1.cell(column=1, row=cells_index).border = thin_border
                self.worksheet_1.cell(column=2, row=cells_index).border = thin_border
                self.worksheet_1.cell(column=3, row=cells_index).border = thin_border
                self.worksheet_1.cell(column=4, row=cells_index).border = thin_border
                self.worksheet_1.cell(column=10, row=cells_index).border = thin_border
                self.worksheet_1.cell(column=11, row=cells_index).border = thin_border
                self.worksheet_1.cell(column=12, row=cells_index).border = thin_border
                self.worksheet_1.cell(column=13, row=cells_index).border = thin_border
                self.worksheet_1.cell(column=14, row=cells_index).border = thin_border
                self.worksheet_1.cell(column=15, row=cells_index).border = thin_border
                self.worksheet_1.cell(column=16, row=cells_index).border = thin_border
                self.worksheet_1.cell(column=17, row=cells_index).border = thin_border
                self.worksheet_1.merge_cells(
                        start_row=cells_index, start_column=1, end_row=cells_index + 4, end_column=1)
                self.worksheet_1.merge_cells(
                        start_row=cells_index, start_column=2, end_row=cells_index + 4, end_column=2)
                self.worksheet_1.merge_cells(
                        start_row=cells_index, start_column=3, end_row=cells_index + 4, end_column=3)
                self.worksheet_1.merge_cells(
                        start_row=cells_index, start_column=4, end_row=cells_index + 4, end_column=4)
                self.worksheet_1.merge_cells(
                        start_row=cells_index, start_column=10, end_row=cells_index + 4, end_column=10)
                self.worksheet_1.merge_cells(
                        start_row=cells_index, start_column=11, end_row=cells_index + 4, end_column=11)
                self.worksheet_1.merge_cells(
                        start_row=cells_index, start_column=12, end_row=cells_index + 4, end_column=12)
                self.worksheet_1.merge_cells(
                        start_row=cells_index, start_column=13, end_row=cells_index + 4, end_column=13)
                self.worksheet_1.merge_cells(
                        start_row=cells_index, start_column=14, end_row=cells_index + 4, end_column=14)
                self.worksheet_1.merge_cells(
                        start_row=cells_index, start_column=15, end_row=cells_index + 4, end_column=15)
                self.worksheet_1.merge_cells(
                        start_row=cells_index, start_column=16, end_row=cells_index + 4, end_column=16)
                self.worksheet_1.merge_cells(
                        start_row=cells_index, start_column=17, end_row=cells_index + 4, end_column=17)
                

                depth_step += 0.5

                cells_index += 5

            # отображаем сведения о слоях
            for layer in layers:
                thin_border = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))
                layer_capacity = layer.depth - prev_depth

                if layer.aquifer == True:
                    aquifer = "Да"
                else:
                    aquifer = "Нет"
                # глубина
                if layer.depth - prev_depth > 0.5:
                    end_row_index = int((layer.depth - prev_depth) / 0.5) * 5 - 1
                    self.worksheet_1.cell(column=6, row=row_index, value=layer.depth)
                    self.worksheet_1.cell(column=7, row=row_index, value=str(layer_capacity))
                    self.worksheet_1.cell(column=9, row=row_index, value=aquifer)
                    self.worksheet_1.cell(column=6, row=row_index).border = thin_border
                    self.worksheet_1.cell(column=6, row=row_index).alignment = Alignment(horizontal='center', vertical='center')
                    self.worksheet_1.merge_cells(
                        start_row=row_index, start_column=6, end_row=row_index + end_row_index, end_column=6)
                    self.worksheet_1.cell(column=7, row=row_index).border = thin_border
                    self.worksheet_1.cell(column=7, row=row_index).alignment = Alignment(horizontal='center', vertical='center')
                    self.worksheet_1.merge_cells(
                        start_row=row_index, start_column=7, end_row=row_index + end_row_index, end_column=7)
                    self.worksheet_1.cell(column=9, row=row_index).border = thin_border
                    self.worksheet_1.cell(column=9, row=row_index).alignment = Alignment(horizontal='center', vertical='center')
                    self.worksheet_1.merge_cells(
                        start_row=row_index, start_column=9, end_row=row_index + end_row_index, end_column=9)
                    row_index += end_row_index + 1
                elif layer.depth - prev_depth == 0.5:
                    self.worksheet_1.cell(column=6, row=row_index, value=layer.depth)
                    self.worksheet_1.cell(column=7, row=row_index, value=str(layer_capacity))
                    self.worksheet_1.cell(column=6, row=row_index).border = thin_border
                    self.worksheet_1.cell(column=6, row=row_index).alignment = Alignment(horizontal='center', vertical='center')
                    self.worksheet_1.merge_cells(
                        start_row=row_index, start_column=6, end_row=row_index + 4, end_column=6)
                    self.worksheet_1.cell(column=7, row=row_index).border = thin_border
                    self.worksheet_1.cell(column=7, row=row_index).alignment = Alignment(horizontal='center', vertical='center')
                    self.worksheet_1.merge_cells(
                        start_row=row_index, start_column=7, end_row=row_index + 4, end_column=7)

                    row_index += 5

                prev_depth = layer.depth

            row_index = 4
            prev_depth = 0
            self.worksheet_2.cell(column=10, row=1, value=well.name)
            for layer in layers:
                thin_border = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))
                self.worksheet_2.cell(column=1, row=row_index, value=prev_depth)
                self.worksheet_2.cell(column=2, row=row_index, value=layer.depth)

                self.worksheet_2.merge_cells(
                        start_row=row_index, start_column=1, end_row=row_index + 4, end_column=1)

                self.worksheet_2.merge_cells(
                        start_row=row_index, start_column=2, end_row=row_index + 4, end_column=2)

                self.worksheet_2.cell(column=3, row=row_index, value=layer.layer_material.name)
                self.worksheet_2.merge_cells(
                        start_row=row_index, start_column=3, end_row=row_index + 4, end_column=10)

                prev_depth = layer.depth
                row_index += 5

            self.workbook.save(filepath)
            self.workbook.close()
        # except Exception as e:
        #     error_message = e
        # finally:
        #     return error_message
